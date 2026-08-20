import datetime
import json

import openpyxl
import pytest

from engine import analyze_file, engine_version, extract_unique_values, preview_file


def test_engine_version_matches_vendored():
    # vendor/wheels/openpyxl-3.1.5-py2.py3-none-any.whl — keep in sync
    assert engine_version() == "3.1.5"


def test_preview_xlsx_single_sheet(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["name", "city", "amount"])
    for i in range(10):
        ws.append([f"Customer {i}", "Berlin", i * 100])
    wb.save(path)

    result = preview_file(str(path))
    assert set(result["sheets"].keys()) == {"Orders"}
    table = result["sheets"]["Orders"]
    assert table["columns"] == ["name", "city", "amount"]
    assert len(table["rows"]) == 5  # max_rows default, not all 10
    assert table["rows"][0] == ["Customer 0", "Berlin", 0]


def test_preview_xlsx_multiple_sheets_are_separate_tables(tmp_path):
    # design.md §5: every sheet of a workbook is its own table
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["order_id", "customer"])
    ws1.append([1, "Anna"])

    ws2 = wb.create_sheet("Customers")
    ws2.append(["customer", "email"])
    ws2.append(["Anna", "anna@example.com"])
    wb.save(path)

    result = preview_file(str(path))
    assert set(result["sheets"].keys()) == {"Orders", "Customers"}
    assert result["sheets"]["Orders"]["columns"] == ["order_id", "customer"]
    assert result["sheets"]["Customers"]["columns"] == ["customer", "email"]


def test_preview_xlsx_respects_max_rows(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a"])
    for i in range(20):
        ws.append([i])
    wb.save(path)

    result = preview_file(str(path), max_rows=3)
    assert len(result["sheets"][ws.title]["rows"]) == 3


def test_preview_xlsx_fewer_rows_than_max(tmp_path):
    path = tmp_path / "small.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(path)

    result = preview_file(str(path))
    table = result["sheets"][ws.title]
    assert table["columns"] == ["a", "b"]
    assert table["rows"] == [[1, 2]]


def test_preview_xlsx_formula_cell_without_cached_value_reads_as_none(tmp_path):
    # openpyxl does not evaluate formulas itself -- if the file was never
    # opened in Excel/Sheets (no cached value), data_only=True yields None
    # rather than the formula or an error. Verified empirically. This
    # sharpens the accepted risk in design.md §10: it is not merely
    # "restore will not bring the formula back" but "a cell with no cached
    # value reads as empty -- silently, with no warning at this level".
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b", "sum"])
    ws.append([5, 10, "=A2+B2"])
    wb.save(path)

    result = preview_file(str(path))
    row = result["sheets"][ws.title]["rows"][0]
    assert row == [5, 10, None]
    assert row[2] != "=A2+B2"  # never show the client a raw formula


def test_preview_csv(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("name,city\nAnna,Moscow\nIvan,Kazan\n", encoding="utf-8")

    result = preview_file(str(path))
    table = result["sheets"]["data"]
    assert table["columns"] == ["name", "city"]
    assert table["rows"] == [["Anna", "Moscow"], ["Ivan", "Kazan"]]


def test_preview_tsv_uses_tab_delimiter(tmp_path):
    path = tmp_path / "data.tsv"
    path.write_text("name\tcity\nAnna\tMoscow\n", encoding="utf-8")

    result = preview_file(str(path))
    table = result["sheets"]["data"]
    assert table["columns"] == ["name", "city"]
    assert table["rows"] == [["Anna", "Moscow"]]


def test_analyze_file_proposes_text_columns_and_leaves_numbers_alone(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer name", "contract amount"])
    for i in range(20):
        ws.append([f"Company {i} Ltd", i * 1000])
    wb.save(path)

    result = analyze_file(str(path))
    table = result["sheets"]["Orders"]
    assert table["columns"] == ["customer name", "contract amount"]
    assert len(table["rows"]) == 5  # preview_rows default, not the whole sample

    guesses = {g["column"]: g for g in table["guesses"]}
    assert guesses["customer name"]["touch"] is True
    assert guesses["contract amount"]["touch"] is False
    # design.md §7: headings are not analyzed, so even a heading that reads
    # exactly like an entity name leaves the entity field empty for the
    # client to fill in
    assert guesses["customer name"]["entity"] is None


def test_analyze_file_sample_bigger_than_preview(tmp_path):
    path = tmp_path / "data.csv"
    # the first 5 rows look numeric but the rest of the sample is text;
    # with a small sample this would be misdetected as numeric
    rows = ["col"] + ["1", "2", "3", "4", "5"] + ["text"] * 10
    path.write_text("\n".join(rows), encoding="utf-8")

    result = analyze_file(str(path), sample_size=20, preview_rows=5)
    guess = result["sheets"]["data"]["guesses"][0]
    assert guess["shape"] == "text"  # the mixed type is only visible in a larger sample


def test_analyze_file_multiple_sheets_analyzed_independently(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["customer", "amount"])
    ws1.append(["Globex LLC", 500])

    ws2 = wb.create_sheet("Staff")
    ws2.append(["employee name", "salary"])
    ws2.append(["John Smith", 80000])
    wb.save(path)

    result = analyze_file(str(path))
    orders_guesses = {g["column"]: g for g in result["sheets"]["Orders"]["guesses"]}
    staff_guesses = {g["column"]: g for g in result["sheets"]["Staff"]["guesses"]}
    # each sheet is shaped independently: text proposed, numbers left alone
    assert orders_guesses["customer"]["touch"] is True
    assert orders_guesses["amount"]["touch"] is False
    assert staff_guesses["employee name"]["touch"] is True
    assert staff_guesses["salary"]["touch"] is False


def test_analyze_file_result_is_json_safe_with_date_column(tmp_path):
    # regression: postMessage in worker.js used to crash on unconverted
    # datetime.datetime values from Excel cells ("[object Object] could not
    # be cloned"). pytest on plain CPython does not exercise that boundary
    # on its own, so this test catches the same class of bug through
    # json.dumps, which has the same type limits as structured clone
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["customer", "date", "Email"])
    for i in range(5):
        ws.append([f"Company {i}", datetime.date(2026, 1, i + 1), f"c{i}@example.com"])
    wb.save(path)

    result = analyze_file(str(path))
    json.dumps(result)  # must not raise TypeError

    table = result["sheets"][ws.title]
    date_guess = next(g for g in table["guesses"] if g["column"] == "date")
    assert date_guess["shape"] == "date"
    assert date_guess["touch"] is False
    assert isinstance(table["rows"][0][1], str)  # date serialized to an ISO string


def test_extract_unique_values_xlsx(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount"])
    for name in ["Anna", "Ivan", "Anna", "Peter", "Anna"]:
        ws.append([name, 100])
    wb.save(path)

    result = extract_unique_values(str(path), {"Orders": ["customer"]})
    table = result["Orders"]["customer"]
    assert set(table["values"]) == {"Anna", "Ivan", "Peter"}
    assert table["unique_count"] == 3
    assert table["truncated"] is False


def test_extract_unique_values_only_requested_columns(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b", "c"])
    ws.append([1, 2, 3])
    wb.save(path)

    result = extract_unique_values(str(path), {"Sheet1": ["a"]})
    assert set(result["Sheet1"].keys()) == {"a"}


def test_extract_unique_values_ignores_sheets_not_requested(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "One"
    ws1.append(["x"])
    ws1.append(["v1"])
    ws2 = wb.create_sheet("Two")
    ws2.append(["y"])
    ws2.append(["v2"])
    wb.save(path)

    result = extract_unique_values(str(path), {"One": ["x"]})
    assert set(result.keys()) == {"One"}


def test_extract_unique_values_truncates_at_max_unique(tmp_path):
    path = tmp_path / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id"])
    for i in range(50):
        ws.append([f"id{i}"])
    wb.save(path)

    result = extract_unique_values(str(path), {"Sheet1": ["id"]}, max_unique=10)
    table = result["Sheet1"]["id"]
    assert table["truncated"] is True
    assert len(table["values"]) <= 10


def test_extract_unique_values_skips_empty(tmp_path):
    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a"])
    ws.append(["x"])
    ws.append([None])
    ws.append([""])
    wb.save(path)

    result = extract_unique_values(str(path), {"Sheet1": ["a"]})
    assert result["Sheet1"]["a"]["values"] == ["x"]


def test_extract_unique_values_csv(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("name\nAnna\nIvan\nAnna\n", encoding="utf-8")

    result = extract_unique_values(str(path), {"data": ["name"]})
    assert set(result["data"]["name"]["values"]) == {"Anna", "Ivan"}


def test_extract_unique_values_csv_temp_path_differs_from_file_pattern(tmp_path):
    # Regression: worker.js writes the CSV to a temp file with a different
    # prefix at every stage (e.g. /tmp/map-7-orders.csv), not the prefix
    # used during preview (/tmp/0-orders.csv). The CSV pseudo sheet name has
    # to match what landed in the config during column mapping -- i.e. it
    # must be derived from file_pattern (the original file name), not from
    # the temp path. Without that, extract_unique_values looks up
    # sheet_columns under the wrong key and silently returns nothing.
    path = tmp_path / "map-7-orders.csv"
    path.write_text("name\nAnna\nIvan\nAnna\n", encoding="utf-8")

    result = extract_unique_values(str(path), {"orders": ["name"]}, file_pattern="orders.csv")
    assert set(result["orders"]["name"]["values"]) == {"Anna", "Ivan"}


def test_preview_file_csv_sheet_name_uses_file_pattern(tmp_path):
    path = tmp_path / "0-orders.csv"
    path.write_text("name\nAnna\n", encoding="utf-8")

    result = preview_file(str(path), file_pattern="orders.csv")
    assert list(result["sheets"].keys()) == ["orders"]


def test_analyze_file_csv_sheet_name_uses_file_pattern(tmp_path):
    path = tmp_path / "5-orders.csv"
    path.write_text("name\nAnna\n", encoding="utf-8")

    result = analyze_file(str(path), file_pattern="orders.csv")
    assert list(result["sheets"].keys()) == ["orders"]


def test_unsupported_extension_rejected(tmp_path):
    path = tmp_path / "data.txt"
    path.write_text("whatever", encoding="utf-8")

    with pytest.raises(ValueError, match="unsupported file format"):
        preview_file(str(path))
