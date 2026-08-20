"""Acceptance criteria (design.md §11) -- gathered in one place.

This does not duplicate coverage that already exists; it points at it and
fills in what was missing:

1. Round-trip `restore(mask(X)) == X` — tests/test_restorer.py
   (test_round_trip_mask_then_restore_recovers_original_xlsx/_csv)
2. Idempotency -- here, below
3. Cross-file token integrity -- here, below (at the mask_file level; the
   build_mapping level is already covered by
   tests/test_mapping.py::test_build_mapping_merges_columns_across_files_into_one_entity)
4. Offline / no network sockets -- blocked globally for the whole test
   run, tests/conftest.py::_block_network_sockets; below, a test proving
   the block actually fires
5. Source values never reach the report -- tests/test_report.py
   (test_report_html_never_contains_touched_original_values)
"""

import socket

import pytest

from conftest import NetworkBlockedError
from mapping import new_mapping, update_entity
from masker import mask_file


def make_mapping(entity, prefix, width, values):
    mapping = new_mapping()
    update_entity(mapping, entity, prefix, width, values)
    return mapping


# ---- 2. idempotency ----


def test_mask_file_idempotent_two_runs_same_input_same_mapping(tmp_path):
    src = tmp_path / "orders.csv"
    src.write_text("name,amount\nAnna,100\nIvan,200\nAnna,50\n", encoding="utf-8")

    config = {
        "version": 1,
        "entities": {
            "customer": {
                "prefix": "CUST",
                "width": 4,
                "columns": [
                    {"file_pattern": "orders.csv", "sheet_name": "orders", "column_name": "name", "extract_regex": None}
                ],
            }
        },
    }
    mapping = make_mapping("customer", "CUST", 4, ["Anna", "Ivan"])

    out1 = tmp_path / "masked1.csv"
    out2 = tmp_path / "masked2.csv"
    summary1 = mask_file(str(src), str(out1), "orders.csv", config, mapping)
    summary2 = mask_file(str(src), str(out2), "orders.csv", config, mapping)

    assert summary1 == summary2
    assert out1.read_text(encoding="utf-8") == out2.read_text(encoding="utf-8")


# ---- 3. cross-file token integrity (mask_file level) ----


def test_mask_file_same_value_same_token_across_two_files(tmp_path):
    orders = tmp_path / "orders.csv"
    orders.write_text("customer,amount\nGlobex LLC,500\n", encoding="utf-8")

    customers = tmp_path / "customers.csv"
    customers.write_text("name,city\nGlobex LLC,Berlin\n", encoding="utf-8")

    config = {
        "version": 1,
        "entities": {
            "customer": {
                "prefix": "CUST",
                "width": 4,
                "columns": [
                    {"file_pattern": "orders.csv", "sheet_name": "orders", "column_name": "customer", "extract_regex": None},
                    {"file_pattern": "customers.csv", "sheet_name": "customers", "column_name": "name", "extract_regex": None},
                ],
            }
        },
    }
    # a single dictionary built from the values of both files -- exactly
    # what build_mapping does in practice (design.md §8)
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])

    out_orders = tmp_path / "masked_orders.csv"
    out_customers = tmp_path / "masked_customers.csv"
    mask_file(str(orders), str(out_orders), "orders.csv", config, mapping)
    mask_file(str(customers), str(out_customers), "customers.csv", config, mapping)

    token_in_orders = out_orders.read_text(encoding="utf-8").strip().split("\n")[1].split(",")[0]
    token_in_customers = out_customers.read_text(encoding="utf-8").strip().split("\n")[1].split(",")[0]
    assert token_in_orders == token_in_customers


# ---- 4. offline / no network sockets ----


def test_network_sockets_are_blocked_during_tests():
    # Prove that _block_network_sockets (conftest.py) actually intercepts a
    # network attempt rather than merely existing as a fixture -- the same
    # "not theory" principle as the rest of the suite.
    with pytest.raises(NetworkBlockedError):
        socket.create_connection(("example.invalid", 80), timeout=1)


# ---- 5. numbers and dates are not modified (v1-beta.md, criterion 5) ----


def test_mask_file_leaves_numbers_and_dates_untouched_xlsx(tmp_path):
    import datetime

    import openpyxl

    src = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount", "invoice_date", "qty", "rate"])
    # rate values are deliberately non-integral: an integral float is a
    # separate case owned by the file format, pinned in the test below
    rows = [
        ("Globex LLC", 120000, datetime.datetime(2026, 1, 15), 3, 0.07),
        ("Acme Inc", -450.75, datetime.datetime(2026, 2, 2), 0, 0.5),
    ]
    for row in rows:
        ws.append(list(row))
    wb.save(src)

    config = {
        "version": 1,
        "entities": {
            "customer": {
                "prefix": "CUST",
                "width": 4,
                "columns": [
                    {
                        "file_pattern": "orders.xlsx",
                        "sheet_name": "Orders",
                        "column_name": "customer",
                        "extract_regex": None,
                    }
                ],
            }
        },
    }
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC", "Acme Inc"])

    out = tmp_path / "masked.xlsx"
    mask_file(str(src), str(out), "orders.xlsx", config, mapping)

    result = list(openpyxl.load_workbook(out, read_only=True).active.iter_rows(values_only=True))
    for original, masked in zip(rows, result[1:]):
        assert masked[0] != original[0]  # the customer column IS masked
        # Everything else comes back unchanged in value -- negative numbers,
        # zero, floats and datetimes are the interesting cases.
        assert masked[1] == original[1]
        assert masked[2] == original[2]
        assert masked[3] == original[3]
        assert masked[4] == original[4]
        # Types survive too, with one exception that belongs to the file
        # format rather than to us: see the test below.
        assert isinstance(masked[2], datetime.datetime)
        assert isinstance(masked[4], float)


def test_xlsx_turns_integral_floats_into_ints_by_itself(tmp_path):
    """Not a defect of ours, and pinned here so nobody mistakes it for one.

    Excel stores every number as a double and openpyxl hands back an int
    whenever that double is integral, so a cell written as 1.0 reads back as
    1. This happens on a plain write/read cycle with no masking involved,
    which is what this test demonstrates. Invariant 5 of v1-beta.md ("types
    are preserved") therefore means the numeric value survives, not that the
    Python type label does.
    """
    import openpyxl

    path = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append([1.0, 0.07])
    wb.save(path)

    row = list(openpyxl.load_workbook(path, read_only=True).active.iter_rows(values_only=True))[0]
    assert row[0] == 1.0 and isinstance(row[0], int)  # integral float -> int
    assert row[1] == 0.07 and isinstance(row[1], float)  # non-integral stays float


def test_mask_file_leaves_numbers_and_dates_untouched_csv(tmp_path):
    # A CSV carries no types, so this is where an over-eager heuristic would
    # tokenize amounts and dates. The masked copy must keep them verbatim.
    src = tmp_path / "orders.csv"
    src.write_text(
        "customer,amount,invoice_date\n"
        "Globex LLC,120000,2026-01-15\n"
        "Acme Inc,-450.75,2026-02-02\n",
        encoding="utf-8",
    )

    config = {
        "version": 1,
        "entities": {
            "customer": {
                "prefix": "CUST",
                "width": 4,
                "columns": [
                    {
                        "file_pattern": "orders.csv",
                        "sheet_name": "orders",
                        "column_name": "customer",
                        "extract_regex": None,
                    }
                ],
            }
        },
    }
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC", "Acme Inc"])

    out = tmp_path / "masked.csv"
    mask_file(str(src), str(out), "orders.csv", config, mapping)

    lines = out.read_text(encoding="utf-8").strip().split("\n")
    assert lines[1].endswith(",120000,2026-01-15")
    assert lines[2].endswith(",-450.75,2026-02-02")
    assert "Globex LLC" not in lines[1]
    assert "Acme Inc" not in lines[2]
