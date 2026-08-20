import openpyxl
import pytest

from config import columns_for_file
from mapping import new_mapping, update_entity
from masker import build_free_text_pattern, mask_file, mask_free_text, mask_scalar


def make_config(file_pattern, sheet_name, entity_columns):
    """entity_columns: {entity: [(column_name, extract_regex_or_None), ...]}"""
    entities = {}
    for entity, cols in entity_columns.items():
        entities[entity] = {
            "prefix": entity[:4].upper(),
            "width": 4,
            "columns": [
                {
                    "file_pattern": file_pattern,
                    "sheet_name": sheet_name,
                    "column_name": name,
                    "extract_regex": regex,
                }
                for name, regex in cols
            ],
        }
    return {"version": 1, "entities": entities}


def make_mapping(entity, prefix, width, values):
    mapping = new_mapping()
    update_entity(mapping, entity, prefix, width, values)
    return mapping


# ---- columns_for_file ----


def test_columns_for_file_groups_by_sheet_and_column():
    config = make_config("orders.xlsx", "Orders", {"customer": [("customer", None)]})
    result = columns_for_file(config, "orders.xlsx")
    assert result == {"Orders": {"customer": {"entity": "customer", "extract_regex": None}}}


def test_columns_for_file_ignores_other_files():
    config = make_config("orders.xlsx", "Orders", {"customer": [("customer", None)]})
    assert columns_for_file(config, "other.xlsx") == {}


# ---- mask_scalar ----


def test_mask_scalar_untouched_column_passthrough():
    mapping = make_mapping("customer", "CUST", 4, ["Acme"])
    assert mask_scalar("Acme", None, None, mapping) == "Acme"


def test_mask_scalar_null_and_empty_preserved_even_if_touched():
    mapping = make_mapping("customer", "CUST", 4, [])
    assert mask_scalar(None, "customer", None, mapping) is None
    assert mask_scalar("", "customer", None, mapping) == ""


def test_mask_scalar_replaces_with_token():
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    result = mask_scalar("Globex LLC", "customer", None, mapping)
    assert result.startswith("CUST-")


def test_mask_scalar_value_not_in_mapping_passthrough():
    mapping = make_mapping("customer", "CUST", 4, ["known"])
    assert mask_scalar("unknown", "customer", None, mapping) == "unknown"


def test_mask_scalar_extract_regex_masks_only_captured_group():
    mapping = make_mapping("product", "PROD", 4, ["4471"])
    token = mapping["entities"]["product"]["values"]["4471"]
    from mapping import format_token

    expected_token = format_token("PROD", token, 4)
    result = mask_scalar("SKU-4471-B", "product", "SKU-(.*?)-[A-Z]", mapping)
    assert result == f"SKU-{expected_token}-B"


def test_mask_scalar_extract_regex_no_match_passthrough():
    mapping = make_mapping("product", "PROD", 4, ["4471"])
    assert mask_scalar("no-match-here", "product", "SKU-(.*?)-[A-Z]", mapping) == "no-match-here"


def test_mask_scalar_extract_regex_captured_value_not_in_mapping_passthrough():
    mapping = make_mapping("product", "PROD", 4, ["9999"])  # not 4471
    result = mask_scalar("SKU-4471-B", "product", "SKU-(.*?)-[A-Z]", mapping)
    assert result == "SKU-4471-B"


# ---- free text ----


def test_free_text_replaces_known_value_embedded_in_sentence():
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    pattern, value_to_token = build_free_text_pattern(mapping)
    result = mask_free_text("Payment from Globex LLC on invoice 45", pattern, value_to_token)
    assert "Globex LLC" not in result
    assert "CUST-" in result
    assert "on invoice 45" in result


def test_free_text_longer_value_wins_over_substring():
    mapping = new_mapping()
    update_entity(mapping, "customer", "CUST", 4, ["Globex LLC", "Globex LLC Plus"])
    pattern, value_to_token = build_free_text_pattern(mapping)
    result = mask_free_text("Customer: Globex LLC Plus", pattern, value_to_token)
    # "Plus" must not be left dangling next to the shorter value's token
    assert "Plus" not in result
    assert value_to_token["Globex LLC Plus"] in result


def test_free_text_no_known_values_passthrough():
    pattern, value_to_token = build_free_text_pattern(new_mapping())
    assert mask_free_text("any text at all", pattern, value_to_token) == "any text at all"


# ---- mask_file: xlsx ----


def test_mask_file_xlsx_end_to_end(tmp_path):
    src = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount", "note"])
    ws.append(["Globex LLC", 500, "Payment from Globex LLC"])
    ws.append(["Acme Inc", 700, "not empty"])
    ws.append(["Globex LLC", 300, "not empty"])  # duplicate value
    wb.save(src)

    config = make_config("orders.xlsx", "Orders", {"customer": [("customer", None)]})
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC", "Acme Inc"])

    out = tmp_path / "masked.xlsx"
    summary = mask_file(str(src), str(out), "orders.xlsx", config, mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    ws2 = result_wb.active
    rows = list(ws2.iter_rows(values_only=True))

    assert rows[0] == ("customer", "amount", "note")  # header row untouched
    assert rows[1][1] == 500  # amount untouched (both type and value)
    assert rows[1][0] == rows[3][0]  # invariant 1: one value -> one token (duplicate)
    assert rows[1][0] != "Globex LLC"  # the customer column is masked
    assert "Globex LLC" not in rows[1][2]  # free text is masked too
    assert summary["cells_masked"] > 0


def test_mask_file_xlsx_trailing_none_does_not_shrink_row(tmp_path):
    # regression: Workbook(write_only=True) silently truncates a row whose
    # last cell is None -- on re-read the row comes back shorter than the
    # original and columns shift. Verified empirically against bare openpyxl
    # before this test was written.
    src = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount", "note"])
    ws.append(["Acme Inc", 700, None])  # note is empty, and it is the last column
    wb.save(src)

    config = make_config("orders.xlsx", "Orders", {"customer": [("customer", None)]})
    mapping = make_mapping("customer", "CUST", 4, ["Acme Inc"])

    out = tmp_path / "masked.xlsx"
    mask_file(str(src), str(out), "orders.xlsx", config, mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    rows = list(result_wb.active.iter_rows(values_only=True))
    assert len(rows[1]) == 3  # not 2 -- the "note" column was not lost


def test_mask_file_xlsx_preserves_row_order_and_untouched_columns(tmp_path):
    src = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "customer"])
    for i in range(10):
        ws.append([i, f"Customer {i}"])
    wb.save(src)

    config = make_config("data.xlsx", "Sheet1", {"customer": [("customer", None)]})
    values = [f"Customer {i}" for i in range(10)]
    mapping = make_mapping("customer", "CUST", 4, values)

    out = tmp_path / "masked.xlsx"
    mask_file(str(src), str(out), "data.xlsx", config, mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    rows = list(result_wb.active.iter_rows(values_only=True))[1:]
    ids = [r[0] for r in rows]
    assert ids == list(range(10))  # row order unchanged


def test_mask_file_xlsx_multiple_sheets_independent(tmp_path):
    src = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["customer"])
    ws1.append(["Globex LLC"])
    ws2 = wb.create_sheet("Customers")
    ws2.append(["name"])
    ws2.append(["Globex LLC"])
    wb.save(src)

    config = make_config("book.xlsx", "Orders", {"customer": [("customer", None)]})
    config["entities"]["customer"]["columns"].append(
        {"file_pattern": "book.xlsx", "sheet_name": "Customers", "column_name": "name", "extract_regex": None}
    )
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])

    out = tmp_path / "masked.xlsx"
    mask_file(str(src), str(out), "book.xlsx", config, mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    v1 = list(result_wb["Orders"].iter_rows(values_only=True))[1][0]
    v2 = list(result_wb["Customers"].iter_rows(values_only=True))[1][0]
    assert v1 == v2  # one entity across both sheets -> one token


def test_mask_file_unsupported_extension(tmp_path):
    src = tmp_path / "data.txt"
    src.write_text("x")
    with pytest.raises(ValueError, match="unsupported file format"):
        mask_file(str(src), str(tmp_path / "out.txt"), "data.txt", make_config("data.txt", "s", {}), new_mapping())


# ---- mask_file: csv ----


def test_mask_file_csv_end_to_end(tmp_path):
    src = tmp_path / "data.csv"
    src.write_text("name,amount\nAnna,100\nIvan,200\nAnna,50\n", encoding="utf-8")

    config = make_config("data.csv", "data", {"customer": [("name", None)]})
    mapping = make_mapping("customer", "CUST", 4, ["Anna", "Ivan"])

    out = tmp_path / "masked.csv"
    mask_file(str(src), str(out), "data.csv", config, mapping)

    lines = out.read_text(encoding="utf-8").strip().split("\n")
    assert lines[0] == "name,amount"
    row1 = lines[1].split(",")
    row3 = lines[3].split(",")
    assert row1[0] == row3[0]  # the same "Anna" -> the same token
    assert row1[1] == "100"  # amount untouched


def test_mask_file_csv_source_path_differs_from_file_pattern(tmp_path):
    # Regression: worker.js reads the CSV from a temp file with its own
    # prefix (e.g. /tmp/mask-src-7-data.csv), different from the original
    # file name ("data.csv") under which the column is recorded in the
    # config (design.md §9.2, found live while verifying §10). The CSV
    # pseudo sheet name must be derived from file_pattern, not source_path,
    # otherwise mask_file finds no column plan for the sheet and "silently"
    # hands the column to the free-text pass instead of the structural one
    # (`mask_row`). For values without extract_regex that is invisible (the
    # free-text pass finds the same whole value anyway -- not a
    # discriminating test), so we test extract_regex specifically: the
    # free-text pass knows nothing about the regex gate and substitutes a
    # token where the structural mask is required to stay silent
    # (design.md §8.1) -- only then is the divergence actually observable.
    src = tmp_path / "mask-src-7-data.csv"
    src.write_text("note\nreference 1234 done\n", encoding="utf-8")

    config = make_config("data.csv", "data", {"order": [("note", r"ID(\d+)")]})
    mapping = make_mapping("order", "ORDR", 6, ["1234"])

    out = tmp_path / "masked.csv"
    summary = mask_file(str(src), str(out), "data.csv", config, mapping)

    lines = out.read_text(encoding="utf-8").strip().split("\n")
    assert lines[1] == "reference 1234 done"  # regex "ID(\d+)" did not match -- left alone
    assert summary["cells_masked"] == 0
