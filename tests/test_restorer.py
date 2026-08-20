import openpyxl

from mapping import new_mapping, token_for, update_entity
from masker import mask_file
from restorer import build_restore_pattern, restore_file, restore_text


def make_mapping(entity, prefix, width, values):
    mapping = new_mapping()
    update_entity(mapping, entity, prefix, width, values)
    return mapping


def make_config(file_pattern, sheet_name, entity_columns):
    entities = {}
    for entity, cols in entity_columns.items():
        entities[entity] = {
            "prefix": entity[:4].upper(),
            "width": 4,
            "columns": [
                {"file_pattern": file_pattern, "sheet_name": sheet_name, "column_name": name, "extract_regex": None}
                for name in cols
            ],
        }
    return {"version": 1, "entities": entities}


# ---- restore_text / build_restore_pattern ----


def test_restore_text_single_token():
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    pattern, token_to_value = build_restore_pattern(mapping)
    token = token_for(mapping, "customer", "Globex LLC")
    assert restore_text(token, pattern, token_to_value) == "Globex LLC"


def test_restore_text_token_embedded_in_sentence():
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    pattern, token_to_value = build_restore_pattern(mapping)
    token = token_for(mapping, "customer", "Globex LLC")
    result = restore_text(f"CUST-0007 underpaid $4,200", pattern, token_to_value)
    # only our own token is replaced, not CUST-0007 (absent from the dict)
    assert result == "CUST-0007 underpaid $4,200"

    result2 = restore_text(f"Payment from {token} on invoice", pattern, token_to_value)
    assert result2 == "Payment from Globex LLC on invoice"


def test_restore_text_multiple_tokens_in_one_string():
    mapping = new_mapping()
    update_entity(mapping, "customer", "CUST", 4, ["Anna", "Ivan"])
    pattern, token_to_value = build_restore_pattern(mapping)
    t1 = token_for(mapping, "customer", "Anna")
    t2 = token_for(mapping, "customer", "Ivan")
    result = restore_text(f"{t1} and {t2} closed a deal", pattern, token_to_value)
    assert result == "Anna and Ivan closed a deal"


def test_restore_text_no_tokens_present_passthrough():
    mapping = make_mapping("customer", "CUST", 4, ["Anna"])
    pattern, token_to_value = build_restore_pattern(mapping)
    assert restore_text("plain text with no tokens", pattern, token_to_value) == "plain text with no tokens"


def test_restore_pattern_empty_mapping():
    pattern, token_to_value = build_restore_pattern(new_mapping())
    assert pattern is None
    assert restore_text("CUST-0001", pattern, token_to_value) == "CUST-0001"


# ---- restore_file: plain text formats ----


def test_restore_file_txt(tmp_path):
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    token = token_for(mapping, "customer", "Globex LLC")
    src = tmp_path / "report.txt"
    src.write_text(f"Customer {token} underpaid 4200", encoding="utf-8")

    out = tmp_path / "report_restored.txt"
    summary = restore_file(str(src), str(out), mapping)

    assert out.read_text(encoding="utf-8") == "Customer Globex LLC underpaid 4200"
    assert summary["tokens_replaced"] == 1


def test_restore_file_json(tmp_path):
    mapping = make_mapping("customer", "CUST", 4, ["Acme Inc"])
    token = token_for(mapping, "customer", "Acme Inc")
    src = tmp_path / "report.json"
    src.write_text(f'{{"customer": "{token}", "amount": 500}}', encoding="utf-8")

    out = tmp_path / "report_restored.json"
    restore_file(str(src), str(out), mapping)

    assert out.read_text(encoding="utf-8") == '{"customer": "Acme Inc", "amount": 500}'


def test_restore_file_md(tmp_path):
    mapping = make_mapping("customer", "CUST", 4, ["Beta LLC"])
    token = token_for(mapping, "customer", "Beta LLC")
    src = tmp_path / "report.md"
    src.write_text(f"# Report\n\n| Customer | Amount |\n|---|---|\n| {token} | 100 |\n", encoding="utf-8")

    out = tmp_path / "report_restored.md"
    restore_file(str(src), str(out), mapping)

    assert "Beta LLC" in out.read_text(encoding="utf-8")
    assert token not in out.read_text(encoding="utf-8")


def test_restore_file_sql(tmp_path):
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    token = token_for(mapping, "customer", "Globex LLC")
    src = tmp_path / "query.sql"
    src.write_text(f"SELECT * FROM orders WHERE customer = '{token}'", encoding="utf-8")

    out = tmp_path / "query_restored.sql"
    restore_file(str(src), str(out), mapping)
    assert "Globex LLC" in out.read_text(encoding="utf-8")


# ---- restore_file: xlsx ----


def test_restore_file_xlsx_cell_based_non_ascii(tmp_path):
    # Deliberately non-ASCII data (accents, a curly apostrophe): plenty of
    # perfectly English business data carries them, and this is the suite's
    # end-to-end check that UTF-8 survives the whole xlsx path -- openpyxl
    # read, token match, write, re-read.
    original = "Zürich Kaffeerösterei O\u2019Brien & Co."
    mapping = make_mapping("customer", "CUST", 4, [original])
    token = token_for(mapping, "customer", original)
    src = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résumé"
    ws.append(["Customer", "Shortfall"])
    ws.append([token, 4200])
    wb.save(src)

    out = tmp_path / "report_restored.xlsx"
    summary = restore_file(str(src), str(out), mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    rows = list(result_wb.active.iter_rows(values_only=True))
    assert rows[1][0] == original
    assert rows[1][1] == 4200  # the number is untouched
    assert summary["cells_restored"] == 1


def test_restore_file_xlsx_works_regardless_of_column_name(tmp_path):
    # design.md §2: restore looks for tokens, not columns -- it works even
    # when the "column" is named differently than during masking, or the
    # token ended up outside its original column (copied into a report by
    # hand, for instance)
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC"])
    token = token_for(mapping, "customer", "Globex LLC")
    src = tmp_path / "custom_report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Arbitrary heading"])
    ws.append([f"See customer {token} in the appendix"])
    wb.save(src)

    out = tmp_path / "custom_report_restored.xlsx"
    restore_file(str(src), str(out), mapping)

    result_wb = openpyxl.load_workbook(out, read_only=True)
    rows = list(result_wb.active.iter_rows(values_only=True))
    assert rows[1][0] == "See customer Globex LLC in the appendix"


# ---- the main acceptance criterion: restore(mask(X)) == X (design.md §11) ----


def test_round_trip_mask_then_restore_recovers_original_xlsx(tmp_path):
    src = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount", "note"])
    ws.append(["Globex LLC", 500, "Payment from Globex LLC on invoice 45"])
    ws.append(["Acme Inc", 700, "no comment"])
    ws.append(["Globex LLC", 300, "repeat customer"])
    wb.save(src)

    config = make_config("orders.xlsx", "Orders", {"customer": ["customer"]})
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC", "Acme Inc"])

    masked_path = tmp_path / "masked.xlsx"
    mask_file(str(src), str(masked_path), "orders.xlsx", config, mapping)

    restored_path = tmp_path / "restored.xlsx"
    restore_file(str(masked_path), str(restored_path), mapping)

    original_wb = openpyxl.load_workbook(src, read_only=True)
    restored_wb = openpyxl.load_workbook(restored_path, read_only=True)
    original_rows = list(original_wb.active.iter_rows(values_only=True))
    restored_rows = list(restored_wb.active.iter_rows(values_only=True))

    assert original_rows == restored_rows


def test_round_trip_mask_then_restore_recovers_original_csv(tmp_path):
    src = tmp_path / "orders.csv"
    src.write_text("name,amount\nAnna,100\nIvan,200\nAnna,50\n", encoding="utf-8")

    config = make_config("orders.csv", "orders", {"customer": ["name"]})
    mapping = make_mapping("customer", "CUST", 4, ["Anna", "Ivan"])

    masked_path = tmp_path / "masked.csv"
    mask_file(str(src), str(masked_path), "orders.csv", config, mapping)

    restored_path = tmp_path / "restored.csv"
    restore_file(str(masked_path), str(restored_path), mapping)

    assert restored_path.read_text(encoding="utf-8") == src.read_text(encoding="utf-8")
