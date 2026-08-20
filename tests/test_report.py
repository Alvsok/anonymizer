import openpyxl

from mapping import new_mapping, update_entity
from report import build_report, collect_column_stats, render_report_html


def make_config(file_pattern, sheet_name, entity_columns):
    entities = {}
    for entity, cols in entity_columns.items():
        entities[entity] = {
            "prefix": entity[:4].upper(),
            "width": 4,
            "columns": [
                {"file_pattern": file_pattern, "sheet_name": sheet_name, "column_name": name, "extract_regex": None}
                for name in cols
            ],
        }
    return {"version": 1, "entities": entities}


def make_mapping(entity, prefix, width, values):
    mapping = new_mapping()
    update_entity(mapping, entity, prefix, width, values)
    return mapping


# ---- collect_column_stats ----


def test_collect_stats_xlsx_touched_and_untouched(tmp_path):
    path = tmp_path / "orders.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["customer", "amount"])
    ws.append(["Globex LLC", 500])
    ws.append(["Acme Inc", 700])
    ws.append(["Globex LLC", 300])  # duplicate -- must not inflate unique_masked
    wb.save(path)

    config = make_config("orders.xlsx", "Orders", {"customer": ["customer"]})
    sheets, row_count = collect_column_stats(str(path), "orders.xlsx", config)

    counterparty = sheets["Orders"]["customer"]
    assert counterparty["touched"] is True
    assert counterparty["entity"] == "customer"
    assert counterparty["unique_masked"] == 2  # Globex LLC, Acme Inc -- not 3

    summa = sheets["Orders"]["amount"]
    assert summa["touched"] is False
    assert set(summa["examples"]) == {500, 700, 300}
    assert row_count == 3


def test_collect_stats_examples_capped_at_max(tmp_path):
    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["note"])
    for i in range(10):
        ws.append([f"value {i}"])
    wb.save(path)

    config = make_config("data.xlsx", "Sheet1", {})
    sheets, _ = collect_column_stats(str(path), "data.xlsx", config, max_examples=3)
    assert len(sheets["Sheet1"]["note"]["examples"]) == 3


def test_collect_stats_empty_values_not_counted_or_shown(tmp_path):
    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b"])
    ws.append(["x", "not empty"])
    ws.append([None, "not empty"])
    ws.append(["x", "not empty"])
    wb.save(path)

    config = make_config("data.xlsx", "Sheet1", {"customer": ["a"]})
    sheets, _ = collect_column_stats(str(path), "data.xlsx", config)
    assert sheets["Sheet1"]["a"]["unique_masked"] == 1  # only "x", not None


def test_collect_stats_csv(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("name,note\nAnna,first\nIvan,second\nAnna,third\n", encoding="utf-8")
    config = make_config("data.csv", "data", {"customer": ["name"]})
    sheets, row_count = collect_column_stats(str(path), "data.csv", config)
    assert sheets["data"]["name"]["unique_masked"] == 2
    assert len(sheets["data"]["note"]["examples"]) == 3
    assert row_count == 3


def test_collect_stats_csv_source_path_differs_from_file_pattern(tmp_path):
    # Regression: for the report worker.js reads the CSV from a temp file
    # with its own prefix (/tmp/report-N-data.csv) that does not match the
    # original name in the config -- see the same case in test_masker.py.
    path = tmp_path / "report-3-data.csv"
    path.write_text("name,note\nAnna,first\n", encoding="utf-8")
    config = make_config("data.csv", "data", {"customer": ["name"]})
    sheets, row_count = collect_column_stats(str(path), "data.csv", config)
    assert sheets["data"]["name"]["touched"] is True
    assert sheets["data"]["name"]["unique_masked"] == 1
    assert row_count == 1


def test_collect_stats_multiple_sheets_independent(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["customer"])
    ws1.append(["A"])
    ws2 = wb.create_sheet("Other")
    ws2.append(["comment"])
    ws2.append(["text"])
    wb.save(path)

    config = make_config("book.xlsx", "Orders", {"customer": ["customer"]})
    sheets, _ = collect_column_stats(str(path), "book.xlsx", config)
    assert sheets["Orders"]["customer"]["touched"] is True
    assert sheets["Other"]["comment"]["touched"] is False


# ---- build_report ----


def test_build_report_aggregates_summary():
    files_stats = {
        "orders.xlsx": {
            "Orders": {
                "customer": {"touched": True, "entity": "customer", "unique_masked": 2},
                "amount": {"touched": False, "entity": None, "examples": [500, 700]},
            }
        }
    }
    config = make_config("orders.xlsx", "Orders", {"customer": ["customer"]})
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC", "Acme Inc"])

    report = build_report(files_stats, {"orders.xlsx": 3}, config, mapping)
    assert report["summary"] == {"files": 1, "rows": 3, "entities": 1, "tokens": 2}
    assert len(report["touched"]) == 1
    assert len(report["untouched"]) == 1


def test_build_report_fingerprint_deterministic():
    config = make_config("a.xlsx", "s", {"customer": ["c"]})
    mapping = make_mapping("customer", "CUST", 4, ["x"])
    r1 = build_report({}, {}, config, mapping)
    r2 = build_report({}, {}, config, mapping)
    assert r1["fingerprint"] == r2["fingerprint"]


def test_build_report_fingerprint_differs_for_different_seed():
    config = make_config("a.xlsx", "s", {"customer": ["c"]})
    m1 = make_mapping("customer", "CUST", 4, ["x"])
    m2 = make_mapping("customer", "CUST", 4, ["x"])  # a fresh seed every time
    r1 = build_report({}, {}, config, m1)
    r2 = build_report({}, {}, config, m2)
    assert r1["fingerprint"] != r2["fingerprint"]


# ---- render_report_html: acceptance criterion design.md §11.5 ----


def test_report_html_never_contains_touched_original_values():
    files_stats = {
        "orders.xlsx": {
            "Orders": {
                "customer": {"touched": True, "entity": "customer", "unique_masked": 1},
            }
        }
    }
    config = make_config("orders.xlsx", "Orders", {"customer": ["customer"]})
    mapping = make_mapping("customer", "CUST", 4, ["Globex LLC Confidential"])

    report = build_report(files_stats, {"orders.xlsx": 1}, config, mapping)
    html = render_report_html(report)
    assert "Globex LLC Confidential" not in html


def test_report_html_shows_untouched_examples():
    files_stats = {
        "orders.xlsx": {
            "Orders": {
                "note": {"touched": False, "entity": None, "examples": ["visible example"]},
            }
        }
    }
    config = make_config("orders.xlsx", "Orders", {})
    mapping = new_mapping()
    report = build_report(files_stats, {"orders.xlsx": 1}, config, mapping)
    html = render_report_html(report)
    assert "visible example" in html


def test_report_html_self_contained_no_external_resources():
    config = make_config("a.xlsx", "s", {})
    html = render_report_html(build_report({}, {}, config, new_mapping()))
    assert "http://" not in html
    assert "https://" not in html
    assert "<script src=" not in html
    assert "<link " not in html  # not a single external stylesheet or font


def test_report_html_escapes_values():
    files_stats = {
        "orders.xlsx": {"Orders": {"a": {"touched": False, "entity": None, "examples": ["<script>alert(1)</script>"]}}}
    }
    config = make_config("orders.xlsx", "Orders", {})
    report = build_report(files_stats, {"orders.xlsx": 1}, config, new_mapping())
    html = render_report_html(report)
    assert "<script>alert(1)</script>" not in html
    assert "&lt;script&gt;" in html
