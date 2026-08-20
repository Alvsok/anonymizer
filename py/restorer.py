import re

import openpyxl

from engine import anchor_trailing_none
from mapping import format_token

TEXT_EXTENSIONS = (".csv", ".tsv", ".md", ".txt", ".json", ".sql")


def build_restore_pattern(mapping):
    """Token -> original value, across all entities at once: `restore`
    looks for token occurrences in text rather than working by column
    (design.md §2). Tokens are globally unique (`config.py` requires unique
    entity prefixes), so one flat dictionary is safe."""
    token_to_value = {}
    for entity in mapping["entities"].values():
        prefix, width = entity["prefix"], entity["width"]
        for value, number in entity["values"].items():
            token_to_value[format_token(prefix, number, width)] = value
    if not token_to_value:
        return None, token_to_value
    ordered = sorted(token_to_value, key=len, reverse=True)
    pattern = re.compile("|".join(re.escape(t) for t in ordered))
    return pattern, token_to_value


def restore_text(text, pattern, token_to_value):
    if pattern is None or not isinstance(text, str) or not text:
        return text
    return pattern.sub(lambda m: str(token_to_value[m.group(0)]), text)


def restore_file(source_path, output_path, mapping):
    """Reads `source_path` and writes a copy with restored values to
    `output_path`. `.xlsx`/`.xls` is processed cell by cell; everything else
    (`.csv`, `.tsv`, `.md`, `.txt`, `.json`, `.sql`, design.md §2) is
    processed as one blob of text, because restore searches for tokens
    rather than structure."""
    if source_path.lower().endswith((".xlsx", ".xls")):
        return _restore_xlsx(source_path, output_path, mapping)
    return _restore_text_file(source_path, output_path, mapping)


def _restore_xlsx(source_path, output_path, mapping):
    pattern, token_to_value = build_restore_pattern(mapping)
    src_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    out_wb = openpyxl.Workbook(write_only=True)
    cells_restored = 0

    for ws in src_wb.worksheets:
        out_ws = out_wb.create_sheet(title=ws.title)
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            restored = [
                restore_text(v, pattern, token_to_value) if isinstance(v, str) else v
                for v in row
            ]
            cells_restored += sum(1 for a, b in zip(row, restored) if a != b)
            out_ws.append(anchor_trailing_none(restored))

    src_wb.close()
    out_wb.save(output_path)
    return {"cells_restored": cells_restored}


def _restore_text_file(source_path, output_path, mapping):
    pattern, token_to_value = build_restore_pattern(mapping)
    with open(source_path, encoding="utf-8") as f:
        text = f.read()
    occurrences = len(pattern.findall(text)) if pattern else 0
    restored = restore_text(text, pattern, token_to_value)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(restored)
    return {"tokens_replaced": occurrences}
