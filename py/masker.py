import re
from pathlib import Path

import openpyxl
import csv as csv_module

from config import columns_for_file
from engine import anchor_trailing_none
from mapping import format_token
from mapping import token_for as mapping_token_for


def _regex_match_span(value_str, pattern):
    """Returns (captured_text, start, end) for group 1, or None when there
    is no match. config.py already guarantees exactly one group during
    config validation."""
    m = re.search(pattern, value_str)
    if not m:
        return None
    return m.group(1), m.start(1), m.end(1)


def mask_scalar(value, entity_name, extract_regex, mapping):
    """A single cell value. `entity_name=None` means the column was not
    mapped, so leave it alone (invariant: untouched columns pass through
    as-is). Emptiness stays emptiness -- invariant 2 (requirements.md §3):
    `None`/`""` are never substituted, even in a mapped column."""
    if entity_name is None:
        return value
    if value is None or value == "":
        return value

    text = str(value)
    if extract_regex:
        match = _regex_match_span(text, extract_regex)
        if not match:
            return value  # regex did not match -- leave it, do not break it
        captured, start, end = match
        token = mapping_token_for(mapping, entity_name, captured)
        if token is None:
            return value  # value is not in the dictionary -- pass it through
        return text[:start] + token + text[end:]

    token = mapping_token_for(mapping, entity_name, text)
    return token if token is not None else value


def build_free_text_pattern(mapping):
    """One compiled regex alternation covering every known value of every
    entity -- design.md §10: "we look for occurrences of values already
    known from the dictionary". Longer values come first so a short
    substring cannot steal the match from a longer one (for example,
    "Acme Holdings Ltd" must not be cut down by "Acme Holdings").
    """
    value_to_token = {}
    for entity in mapping["entities"].values():
        prefix, width = entity["prefix"], entity["width"]
        for value, number in entity["values"].items():
            if isinstance(value, str) and value:
                value_to_token[value] = format_token(prefix, number, width)
    if not value_to_token:
        return None, value_to_token
    ordered = sorted(value_to_token, key=len, reverse=True)
    pattern = re.compile("|".join(re.escape(v) for v in ordered))
    return pattern, value_to_token


def mask_free_text(text, pattern, value_to_token):
    if pattern is None or not isinstance(text, str) or not text:
        return text
    return pattern.sub(lambda m: value_to_token[m.group(0)], text)


def mask_row(row, column_plan_by_index, mapping, free_text_pattern, value_to_token):
    """column_plan_by_index: {column_index: {"entity":..., "extract_regex":...}}."""
    result = []
    for i, value in enumerate(row):
        plan = column_plan_by_index.get(i)
        if plan:
            result.append(mask_scalar(value, plan["entity"], plan["extract_regex"], mapping))
        elif isinstance(value, str):
            result.append(mask_free_text(value, free_text_pattern, value_to_token))
        else:
            result.append(value)
    return result


def _column_plan_by_index(header_row, sheet_plan_by_name):
    plan = {}
    for i, name in enumerate(header_row):
        name = str(name) if name is not None else f"col_{i}"
        if name in sheet_plan_by_name:
            plan[i] = sheet_plan_by_name[name]
    return plan


def mask_file(source_path, output_path, file_pattern, config, mapping):
    """Reads `source_path` and writes the masked copy to `output_path` in
    the same format. `config`: an AnonConfig dict (`config.py`).
    `mapping`: a finished token dictionary (`mapping.py`). Returns a summary
    `{"cells_masked": int}`.
    """
    if source_path.lower().endswith((".xlsx", ".xls")):
        return _mask_xlsx(source_path, output_path, file_pattern, config, mapping)
    if source_path.lower().endswith((".csv", ".tsv")):
        return _mask_csv(source_path, output_path, file_pattern, config, mapping)
    raise ValueError(f"unsupported file format: {source_path}")


def _mask_xlsx(source_path, output_path, file_pattern, config, mapping):
    plan_by_sheet = columns_for_file(config, file_pattern)
    free_text_pattern, value_to_token = build_free_text_pattern(mapping)

    src_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    out_wb = openpyxl.Workbook(write_only=True)
    cells_masked = 0

    for ws in src_wb.worksheets:
        sheet_plan = plan_by_sheet.get(ws.title, {})
        out_ws = out_wb.create_sheet(title=ws.title)
        header = None
        plan_by_index = {}
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            if header is None:
                header = row
                plan_by_index = _column_plan_by_index(header, sheet_plan)
                out_ws.append(anchor_trailing_none(row))  # header row is never masked
                continue
            masked = mask_row(row, plan_by_index, mapping, free_text_pattern, value_to_token)
            cells_masked += sum(1 for a, b in zip(row, masked) if a != b)
            out_ws.append(anchor_trailing_none(masked))

    src_wb.close()
    out_wb.save(output_path)
    return {"cells_masked": cells_masked}


def _mask_csv(source_path, output_path, file_pattern, config, mapping):
    # file_pattern -- the original file name, not source_path (the worker's
    # temp path, which carries a stage-specific prefix) -- otherwise the CSV
    # pseudo sheet name will not match what the config stored during column
    # mapping (design.md §9.2, found live while verifying §10).
    sheet_name = Path(file_pattern).stem
    plan_by_sheet = columns_for_file(config, file_pattern)
    sheet_plan = plan_by_sheet.get(sheet_name, {})
    free_text_pattern, value_to_token = build_free_text_pattern(mapping)
    delimiter = "\t" if source_path.lower().endswith(".tsv") else ","
    cells_masked = 0

    with open(source_path, newline="", encoding="utf-8") as fin, open(
        output_path, "w", newline="", encoding="utf-8"
    ) as fout:
        reader = csv_module.reader(fin, delimiter=delimiter)
        writer = csv_module.writer(fout, delimiter=delimiter)
        header = next(reader, None)
        if header is None:
            return {"cells_masked": 0}
        writer.writerow(header)
        plan_by_index = _column_plan_by_index(header, sheet_plan)
        for row in reader:
            masked = mask_row(row, plan_by_index, mapping, free_text_pattern, value_to_token)
            cells_masked += sum(1 for a, b in zip(row, masked) if a != b)
            writer.writerow(masked)

    return {"cells_masked": cells_masked}
