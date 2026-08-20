import csv
import datetime
from pathlib import Path

import openpyxl

from heuristics import guess_column


def engine_version():
    return openpyxl.__version__


SUPPORTED_INPUT_FORMATS = (".xlsx", ".csv", ".tsv")


def unsupported_format_message(path):
    """What the client sees when the engine cannot read the file they picked.

    `.xls` gets its own wording because it is the one rejection people find
    surprising: it *is* Excel, just the pre-2007 binary format, and openpyxl
    reads only the modern zip-based `.xlsx`. Writing `.xls` is not possible
    either, and masking has to return the same format it was given, so
    supporting it is a separate project rather than a fix
    (`v1-beta.md`, "Форматы").
    """
    if path.lower().endswith(".xls"):
        return (
            "the old .xls format is not supported -- open the file in Excel "
            "or LibreOffice, save it as .xlsx and try again"
        )
    return (
        f"unsupported file format: {path} "
        f"(supported: {', '.join(SUPPORTED_INPUT_FORMATS)})"
    )


def _json_safe_value(value):
    """Coerce a cell value to a type that survives the Pyodide -> JS
    boundary (structured clone via postMessage). Without this a
    `datetime.datetime` from an Excel cell, for one, stays a PyProxy and
    breaks `postMessage` in worker.js -- verified empirically; tests on
    plain CPython do not see that boundary at all."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
        return value.isoformat()
    return str(value)


def anchor_trailing_none(row):
    """`openpyxl.Workbook(write_only=True)` silently truncates a row when
    its last cell is `None` (verified empirically: adjacent `None`s are
    fine, only the very last cell of the row decides this). Without this
    fix a re-read row comes back shorter than the original and columns
    shift for any index-based reader. `None` and `""` are indistinguishable
    when openpyxl reads the file anyway (see `design.md` §10), so swapping
    the last cell for `""` changes nothing for the reader and only fixes
    the row length. Shared helper -- `masker.py`, `restorer.py` and anyone
    else writing xlsx through write_only needs it."""
    if row and row[-1] is None:
        row = row[:-1] + [""]
    return row


def preview_file(path, file_pattern=None, max_rows=5):
    """Every sheet of a workbook is its own table (design.md §5). For
    CSV/TSV, which have no sheets, the single table is named after the file
    without its extension, so the result has the same shape as for xlsx.

    `file_pattern`: the original file name, used for the CSV pseudo sheet
    name instead of `path`. In the calling code (`worker.js`) `path` is the
    worker's temp file, carrying a different prefix at each stage
    (preview/dictionary/masking/report), so it is not a stable identifier
    for the same file across stages. `None` keeps backward compatibility:
    the pseudo sheet name is taken from `path`, as before.
    """
    if path.lower().endswith(".xlsx"):
        return _preview_xlsx(path, max_rows)
    if path.lower().endswith((".csv", ".tsv")):
        return _preview_csv(path, file_pattern, max_rows)
    raise ValueError(unsupported_format_message(path))


def _preview_xlsx(path, max_rows):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        columns = None
        rows = []
        for row in ws.iter_rows(values_only=True):
            if columns is None:
                columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                continue
            rows.append(list(row))
            if len(rows) >= max_rows:
                break
        sheets[ws.title] = {"columns": columns or [], "rows": rows}
    wb.close()
    return {"sheets": sheets}


def _preview_csv(path, file_pattern, max_rows):
    sheet_name = Path(file_pattern or path).stem
    delimiter = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        columns = next(reader, [])
        rows = []
        for row in reader:
            rows.append(row)
            if len(rows) >= max_rows:
                break
    return {"sheets": {sheet_name: {"columns": columns, "rows": rows}}}


def extract_unique_values(path, sheet_columns, file_pattern=None, max_unique=100_000):
    """Streaming projection (design.md §9.2): one pass over the file,
    keeping only the distinct values of the requested columns -- not every
    row, not every column, only what link detection asks for
    (`sheet_columns`: {sheet_name: [column_name, ...]}).

    Once a column reaches `max_unique` distinct values its projection stops
    (`truncated=True`) -- we use whatever was collected and mark the
    estimate as approximate further up the stack (§5).

    `file_pattern`: see `preview_file` -- gives the CSV pseudo sheet name a
    stable value across worker.js stages; `None` keeps backward
    compatibility (name taken from `path`, as before).
    """
    if path.lower().endswith(".xlsx"):
        return _extract_unique_xlsx(path, sheet_columns, max_unique)
    if path.lower().endswith((".csv", ".tsv")):
        return _extract_unique_csv(path, sheet_columns, file_pattern, max_unique)
    raise ValueError(unsupported_format_message(path))


def _extract_unique_xlsx(path, sheet_columns, max_unique):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for ws in wb.worksheets:
        wanted = sheet_columns.get(ws.title)
        if not wanted:
            continue
        col_index = {}
        sets = {name: set() for name in wanted}
        truncated = {name: False for name in wanted}
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not header_seen:
                for j, c in enumerate(row):
                    name = str(c) if c is not None else f"col_{j}"
                    if name in wanted:
                        col_index[name] = j
                header_seen = True
                continue
            for name, idx in col_index.items():
                if truncated[name] or idx >= len(row):
                    continue
                v = row[idx]
                if v is None or v == "":
                    continue
                v = _json_safe_value(v)
                if v not in sets[name] and len(sets[name]) >= max_unique:
                    truncated[name] = True
                    continue
                sets[name].add(v)
        result[ws.title] = {
            name: {
                "values": sorted(sets[name], key=str),
                "unique_count": len(sets[name]),
                "truncated": truncated[name],
            }
            for name in wanted
        }
    wb.close()
    return result


def _extract_unique_csv(path, sheet_columns, file_pattern, max_unique):
    sheet_name = Path(file_pattern or path).stem
    wanted = sheet_columns.get(sheet_name)
    if not wanted:
        return {}
    delimiter = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        header = next(reader, [])
        col_index = {name: j for j, name in enumerate(header) if name in wanted}
        sets = {name: set() for name in wanted}
        truncated = {name: False for name in wanted}
        for row in reader:
            for name, idx in col_index.items():
                if truncated[name] or idx >= len(row):
                    continue
                v = row[idx]
                if v is None or v == "":
                    continue
                if v not in sets[name] and len(sets[name]) >= max_unique:
                    truncated[name] = True
                    continue
                sets[name].add(v)
    return {
        sheet_name: {
            name: {
                "values": sorted(sets[name], key=str),
                "unique_count": len(sets[name]),
                "truncated": truncated[name],
            }
            for name in wanted
        }
    }


def analyze_file(path, file_pattern=None, sample_size=50, preview_rows=5):
    """Level 0 (design.md §7): a preview for the UI plus a per-column guess,
    computed separately for each sheet (design.md §5).

    The guess is computed over a larger sample (`sample_size`) than what is
    shown to the user (`preview_rows`): the sample is cheap (not the whole
    file) but bigger than five rows, otherwise the shape of the values is
    detected unreliably. `guess_column` sees the original types (dates as
    `date`/`datetime`) -- the JSON-safe coercion is applied only to the rows
    that actually cross into JS, not to the ones feeding the heuristic.

    `file_pattern`: see `preview_file` -- a CSV pseudo sheet name that does
    not depend on the worker's temp path.
    """
    sample = preview_file(path, file_pattern=file_pattern, max_rows=sample_size)
    result_sheets = {}
    for sheet_name, table in sample["sheets"].items():
        columns = table["columns"]
        guesses = []
        for i, name in enumerate(columns):
            values = [row[i] for row in table["rows"] if i < len(row)]
            guesses.append({"column": name, **guess_column(values)})
        safe_rows = [
            [_json_safe_value(v) for v in row] for row in table["rows"][:preview_rows]
        ]
        result_sheets[sheet_name] = {"columns": columns, "rows": safe_rows, "guesses": guesses}
    return {"sheets": result_sheets}
